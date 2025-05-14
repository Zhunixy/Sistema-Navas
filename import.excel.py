import openpyxl
import mysql.connector
from mysql.connector import Error
import os
from pathlib import Path
from dotenv import load_dotenv
import re

# Carrega variáveis do .env
load_dotenv()

# Configuração do MySQL
DB_CONFIG = {
    'host': os.getenv('DB_HOST'),
    'database': os.getenv('DB_NAME'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD')
}

def limpar_preco(valor):
    """Converte valores de preço para float"""
    if isinstance(valor, str):
        valor = valor.replace('R$', '').replace('$', '').strip().replace(',', '.')
    try:
        return float(valor)
    except:
        return 0.0

def limpar_codigo(codigo):
    """Remove pontos e formata o código"""
    if isinstance(codigo, str):
        # Remove espaços e pontos
        codigo = codigo.strip().replace('.', '').replace(',', '')
        # Pega apenas o primeiro código se houver múltiplos separados por vírgula
        if ',' in codigo:
            codigo = codigo.split(',')[0].strip()
        return codigo
    elif isinstance(codigo, (int, float)):
        return str(int(codigo))
    return str(codigo)

def extrair_codigos_alternativos(texto):
    """Extrai códigos alternativos do texto de observações"""
    if not texto:
        return []
    
    codigos = []
    
    # Padrões para encontrar códigos alternativos
    padroes = [
        r'Códigos? alternativos?:?\s*([\d\s,\.]+)',
        r'Códigos?:?\s*([\d\s,\.]+)',
        r'Outros códigos?:?\s*([\d\s,\.]+)',
        r'Também conhecido como:?\s*([\d\s,\.]+)'
    ]
    
    for padrao in padroes:
        matches = re.search(padrao, texto, re.IGNORECASE)
        if matches:
            codigos_str = matches.group(1)
            # Separa por vírgulas, pontos e vírgulas ou espaços
            for codigo in re.split(r'[,;\s]+', codigos_str):
                codigo_limpo = codigo.strip().replace('.', '').replace(',', '')
                if codigo_limpo and codigo_limpo.isdigit():
                    codigos.append(codigo_limpo)
    
    return list(set(codigos))  # Remove duplicados

def importar_excel_para_mysql(caminho_arquivo):
    conexao = None
    try:
        # Verifica se o arquivo existe
        if not os.path.exists(caminho_arquivo):
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")

        # Carrega a planilha Excel
        planilha = openpyxl.load_workbook(caminho_arquivo)
        sheet = planilha.active
        
        # Conecta ao MySQL
        conexao = mysql.connector.connect(**DB_CONFIG)
        cursor = conexao.cursor()
        
        # Limpa as tabelas antes de importar
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        cursor.execute("TRUNCATE TABLE produto_codigos")
        cursor.execute("TRUNCATE TABLE produtos")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        
        # Lê cada linha (ignorando cabeçalho)
        for linha in sheet.iter_rows(min_row=2, values_only=True):
            try:
                codigo_principal = limpar_codigo(linha[0])  # Coluna A - Código principal
                descricao = str(linha[1]).strip() if linha[1] else ''  # Coluna B - Produto
                fabricante = str(linha[2]).strip() if len(linha) > 2 and linha[2] else 'APEX'  # Coluna C - FAB
                observacoes = str(linha[3]).strip() if len(linha) > 3 and linha[3] else ''  # Coluna D - OBSERVAÇÕES
                preco_sp = limpar_preco(linha[4]) if len(linha) > 4 else 0.0  # Coluna E - PROMO SP
                preco_mg = limpar_preco(linha[5]) if len(linha) > 5 else 0.0  # Coluna F - PROMO MG

                if not codigo_principal or not descricao:
                    print(f"⚠️ Linha ignorada (dados faltantes): {linha}")
                    continue
                
                # Padrão de nome de imagem
                imagem = f"{codigo_principal}.jpg"
                
                # Insere o produto principal
                sql_produto = """INSERT INTO produtos 
                         (descricao, fabricante, observacoes, preco_sp, preco_mg, imagem_path) 
                         VALUES (%s, %s, %s, %s, %s, %s)"""
                valores_produto = (descricao, fabricante, observacoes, preco_sp, preco_mg, imagem)
                
                cursor.execute(sql_produto, valores_produto)
                produto_id = cursor.lastrowid
                
                # Insere o código principal
                sql_codigo = """INSERT INTO produto_codigos 
                         (produto_id, codigo, principal) 
                         VALUES (%s, %s, %s)"""
                cursor.execute(sql_codigo, (produto_id, codigo_principal, 1))
                
                # Extrai e insere códigos alternativos das observações
                codigos_alternativos = extrair_codigos_alternativos(observacoes)
                for codigo in codigos_alternativos:
                    if codigo != codigo_principal:  # Evita duplicar o código principal
                        try:
                            cursor.execute(sql_codigo, (produto_id, codigo, 0))
                        except mysql.connector.IntegrityError as e:
                            print(f"⚠️ Código duplicado ignorado: {codigo} para produto {codigo_principal}")
                            continue
                
                print(f"✓ Adicionado: {codigo_principal} - {descricao} | Códigos alternativos: {len(codigos_alternativos)}")
                
            except Exception as e:
                print(f"⚠️ Erro na linha {linha}: {str(e)}")
                continue
        
        conexao.commit()
        print(f"\n✅ Importação concluída! {sheet.max_row - 1} linhas processadas.")
        
    except FileNotFoundError as e:
        print(f"❌ Erro: {e}")
        print("Certifique-se que:")
        print(f"1. O arquivo '{caminho_arquivo}' está na pasta correta")
        print(f"2. O nome do arquivo está exatamente igual (incluindo maiúsculas/minúsculas)")
        print(f"3. A extensão do arquivo é .xlsx")
    except Error as e:
        print(f"❌ Erro MySQL: {e}")
    except Exception as e:
        print(f"❌ Erro inesperado: {e}")
    finally:
        if conexao and conexao.is_connected():
            cursor.close()
            conexao.close()
        
# Uso
if __name__ == "__main__":
    # Caminho para o arquivo Excel (coloque na mesma pasta ou especifique o caminho completo)
    caminho_excel = "PROMOCAO APEX FINALIZADA.xlsx"
    
    # Verifica o caminho absoluto (para debug)
    caminho_absoluto = os.path.abspath(caminho_excel)
    print(f"Procurando arquivo em: {caminho_absoluto}")
    
    importar_excel_para_mysql(caminho_excel)